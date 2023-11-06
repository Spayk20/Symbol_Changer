# импортирование модулей python
import tkinter as tk
from tkinter import filedialog, RAISED, BOTH, DISABLED, ACTIVE
import tkinter.messagebox as msg
import os

import openpyxl
import pandas as pd
from regions import *


# класс главного окна
class Main(tk.Tk):
    items = []
    def __init__(self):
        super().__init__()

        self.title('Умлаут мастер')
        self.geometry('560x200')
        self.resizable(height=False, width=False)
        self.number_y = 0

        self.main_frm = tk.Frame(self, relief=RAISED, borderwidth=2)
        self.main_frm.pack(fill=BOTH, expand=True)

        self.frame1 = tk.Frame(self.main_frm)
        self.frame1.place(relx=0, rely=0)

        self.frame2 = tk.Frame(self.main_frm)
        self.frame2.place(relx=0, rely=0.2)

        self.frame3 = tk.Frame(self.main_frm)
        self.frame3.place(relx=0, rely=0.4)

        self.frame_inf = tk.Frame(self.main_frm, relief=RAISED, borderwidth=1, bg='black')
        self.frame_inf.place(relx=0.45, rely=0, width=305, height=155)

        self.info_lbl = tk.Label(self.frame_inf, text=f">>> Загрузите файл!", fg="white", bg='black')
        self.info_lbl.place(relx=0, rely=0)

        self.listbox = tk.Listbox(self.frame_inf, fg="green", bg='black', bd=0, relief=tk.FLAT, highlightthickness=0)
        self.listbox.place(relx=0, rely=0.15, width=305, height=155)

        self.open_btn = tk.Button(self.frame1, text='Открыть файл', command=self.file_open)
        self.open_btn.pack(side=tk.LEFT)

        self.var = tk.IntVar()
        self.var.set(0)
        self.ww_btn = tk.Radiobutton(self.frame2, text='Весь мир', variable=self.var, value=0)
        self.ww_btn.pack(side=tk.LEFT)

        self.dach_btn = tk.Radiobutton(self.frame2, text='DACH', variable=self.var, value=1)
        self.dach_btn.pack(side=tk.LEFT)

        self.conv_btn = tk.Button(self.frame3, text='Преобразовать', command=self.conv_xls_to_csv)
        self.conv_btn.config(state=DISABLED)
        self.conv_btn.pack(side=tk.LEFT)

        self.closeButton = tk.Button(self, text="Закрыть", command=self.quit)
        self.closeButton.pack(side=tk.RIGHT, padx=5, pady=5)

    def change_state(self):

        if self.conv_btn['state'] == self.NORMAL:
            self.conv_btn['state'] = self.DISABLED
        else:
            self.conv_btn['state'] = self.NORMAL

    def file_open(self):
        self.xls_file = filedialog.askopenfilename()

        while self.xls_file and not self.xls_file.endswith(".xls"):
            msg.showerror("Ошибка", "Пожалуйста, выберите .xls файл")
            self.xls_file = filedialog.askopenfilename()

        if self.xls_file:
            self.info_lbl["text"] = f">>> {os.path.basename(self.xls_file)} Загружен!"
            self.info_lbl["fg"] = "green"
            self.conv_btn.config(state=ACTIVE)
            self.listbox.delete(0, 4)

    def conv_xls_to_csv(self):
        # Конвертация файла в .xlsx
        opened_file = self.xls_file
        file_name = os.path.splitext(os.path.basename(opened_file))[0]

        read_file = pd.read_excel(opened_file, header=None)
        read_file.to_excel(f'{file_name}.xlsx', index=False, header=False)

        # Открытие в openpyxl
        book = openpyxl.open(filename=f"{file_name}.xlsx")
        sheet = book.active

        sheet.delete_rows(idx=1, amount=1)
        sheet.delete_rows(idx=sheet.max_row, amount=1)
        sheet.delete_rows(idx=sheet.max_row, amount=1)
        sheet.delete_rows(idx=sheet.max_row, amount=1)

        colls = (2, 3, 4)  # столбцы для обработки

        for j in colls:
            self.changes = 0
            ch = 0
            for i in range(2, sheet.max_row + 1):
                cell_val = sheet.cell(row=i, column=j).value

                if self.var.get() == 0:
                    val, ch = de_to_ascii(cell_val, ww_sym)

                if self.var.get() == 1:
                    val, ch = de_to_ascii(cell_val, de_sym)

                self.changes += ch
                sheet.cell(row=i, column=j, value=val)

            self.listbox.insert(tk.END, f">>> Заменено {self.changes} символов в столбце {sheet.cell(row=1, column=j).value}")
            print(f'Заменено {self.changes} символов в столбце {sheet.cell(row=1, column=j).value}')

        book.save(f'RESULT_{file_name}.xlsx')

        """
        Emails parsing from secondary to primary
        """
        for i in range(2, sheet.max_row + 1):
            primary_email = sheet.cell(row=i, column=5).value
            secondary_email = sheet.cell(row=i, column=6).value

            if (primary_email is None) or (primary_email == ''):
                primary_email = secondary_email
            sheet.cell(row=i, column=5, value=primary_email)

        print('Столбец Secondary email удален')
        self.listbox.insert(tk.END, ">>> Столбец Secondary email удален")
        sheet.delete_cols(6), book.save(f'RESULT_{file_name}.xlsx')

        os.remove(f'{file_name}.xlsx')
        conv_file = pd.read_excel(f'RESULT_{file_name}.xlsx')
        conv_file.to_csv(f'RESULT_{file_name}.csv', index=False, header=True)
        self.listbox.insert( tk.END, ">>> Готово! :)")

    def quit(self):
        self.destroy()


# создание окна
# запуск окна
if __name__ == "__main__":
    window = Main()
    window.mainloop()

